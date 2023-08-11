from enum import Enum
import os
import yaml
from openpyxl import load_workbook

# class TreePart(Enum):
#     BRANCH = 1
#     LEAF = 2

# eel specific
# class FileType(Enum):
#     OBJECT = 1
#     CONFIG = 2

# class FileSubType(Enum):
#     OBJECT = 1
#     CONFIG_EXPLICIT = 2
#     CONFIG_VIRTUAL = 3
#     CONFIG_MIXED = 4

# class FilePartSubType(Enum):
#     OBJECT = 1
#     CONFIG_EXPLICIT = 2
#     CONFIG_VIRTUAL = 3

class GenericTree:
    def add_node(self, node, parent=None):
        if not parent is None:
            parent.add_child(node)
            self.add_index(node)
        else:
            self.root = node
            self.index = {}
            self.index['.'] = self.root
        return node
    
    def add_index(self, node):
        path = node.get_file_path()
        if path in self.index:
            print('Warning, duplicate data path found, index will reference most recent: ' + path)
        self.index[path] = node

    def remove_node(self, node):
        if node.parent_folder:
            node.parent_folder.remove_child(node)
        else:
            self.root.remove_child(node)
        del self.index[node.get_file_path()]

    def get_node_by_path(self, file_path):
        return self.index.get(file_path)

    def display_tree(self):
        self.root.display()

class FileTree(GenericTree):
    # def __init__(self, path, config_ext='.eel.yml'):
    def __init__(self, path):
        if os.path.isdir(path):
            self.root = Folder(path)
        elif os.path.isfile(path):
            self.root = File(path)
        # else: #should be file part, edge case not tested
        #     self.root = FilePart(path)
        
        # self.config_ext = config_ext

        self.index = {}
        self.index['.'] = self.root

    
    def add_folder(self, name, parent=None):
        folder = Folder(name, parent)
        return self.add_node(folder, parent)

    def add_file(self, name, parent=None):
        file = File(name, parent)
        # if self.is_config_file(file.get_file_path(absolute=True)):
        #     file._file_type = FileType.CONFIG
        return self.add_node(file, parent)

    def add_file_part(self, name, parent=None, content=None):
        file_part = FilePart(name, parent, content)
        return self.add_node(file_part, parent)

class LeafNodeMixin:
    def __init__(self):
        # self.tree_part = TreePart.LEAF
        pass

    def display(self):
        print(self.get_file_path())

class BranchNodeMixin:
    def __init__(self):
        # self.tree_part = TreePart.BRANCH
        self.children = {} 

    def display(self):
        LeafNodeMixin.display(self)
        for child in self.children.values():
            child.display()

    def add_child(self, child_node):
        self.children[child_node.name] = child_node  

    def remove_child(self, child_node):
        if child_node.name in self.children:
            del self.children[child_node.name]  


class GenericNode:
    def __init__(self, name, parent=None):
        self.name = name
        self.parent = parent
        
    def get_file_path(self, absolute=False):
        path_parts = []
        node = self
        while node is not None:
            path_parts.insert(0, node.name)
            node = node.parent
        if not absolute:
            path_parts[0] = '.'
        file_path = os.path.normpath("/".join(path_parts))
        return file_path
    
    @property
    def file_extension(self):
        node = self
        file_extension = None
        while not isinstance(node, Folder):
            if isinstance(node, File):
                file_extension =  node.name.split(".")[-1]
                break
            else:
                node = node.parent
        return file_extension
    
    @property
    def siblings(self):
        return self.parent.children    

class Folder(GenericNode, BranchNodeMixin):
    def __init__(self, name, parent=None):
        super().__init__(name, parent)
        BranchNodeMixin.__init__(self)
    
class File(GenericNode, BranchNodeMixin):
    # def __init__(self, name, parent=None, file_type = FileType.OBJECT):
    def __init__(self, name, parent=None):
        super().__init__(name, parent)
        BranchNodeMixin.__init__(self)
        # self._file_type = file_type
    
class FilePart(GenericNode, LeafNodeMixin):
    def __init__(self, name, parent, content):
        super().__init__(name, parent)
        LeafNodeMixin.__init__(self)
        self.content = content


def get_worksheets_list(file_path):
    workbook = load_workbook(filename=file_path, read_only=True)
    worksheet_names = workbook.sheetnames
    workbook.close()
    return worksheet_names

def has_multiple_documents(file_path):
    with open(file_path, 'r') as file:
        yaml_text = file.read()
        documents = list(yaml.safe_load_all(yaml_text))
        return len(documents) > 1
    
def get_document_dict(file_path):
    with open(file_path, 'r') as file:
        yaml_text = file.read()
        documents = list(yaml.safe_load_all(yaml_text))
        return documents

def populate_tree(file_tree, basepath, parent=None):
    if os.path.isdir(basepath):
        if parent is None:
            folder = file_tree.root
        else:
            folder = file_tree.add_folder(os.path.basename(basepath), parent)

        for item in os.listdir(basepath):
            file_path = os.path.join(basepath, item)
            populate_tree(file_tree, file_path, parent=folder)

    else:
        file = file_tree.add_file(os.path.basename(basepath), parent)

        file_path = file.get_file_path(absolute=True)
        if file.file_extension == 'xlsx':
            worksheets = get_worksheets_list(file_path)
            for ws in worksheets:
                file_tree.add_file_part(ws, file)
        elif file.file_extension == 'yml':
            documents = get_document_dict(file_path)
            for i in range(len(documents)):
                file_tree.add_file_part(str(i), file, documents[i])
        else:
            file_tree.add_file_part("0",file)
            
            # if len(documents) > 1:
            #     for doc in documents:
            #         # assumes dict is ordered a la python 3.7+
            #         # TODO: change to sub-path
            #         _, first_value = next(iter(doc.items()))
            #         file_tree.add_file_part(first_value, file)

def merge_dicts_by_top_level_keys(*dicts):
    merged_dict = {}
    for dict_ in dicts:
        for key, value in dict_.items():
            if key in merged_dict:
                # Perform merge operation on the value of the existing key
                if isinstance(value, dict) and isinstance(merged_dict[key], dict):
                    merged_dict[key].update(value)
            else:
                # Add a new key-value pair to the merged dictionary
                merged_dict[key] = value
    return merged_dict

def find_replace_values(dictionary, find_replace_dict):
    for key, value in dictionary.items():
        if isinstance(value, dict):
            find_replace_values(dictionary[key], find_replace_dict)
        elif value in find_replace_dict:
            dictionary[key] = find_replace_dict[value]

class ConfigMixin():
    def __init__(self,config_target: GenericNode,  config_source: FilePart): 

        self.container_name = self.get_container_name(config_target)
        if config_source:
            self.config_explicit = config_source.content
        else:
            self.config_explicit = None
        # print(self.config_explicit)

    def get_container_name(self, _from_target=None):
        if _from_target:
            return _from_target.name
        # elif _from_source:
        #     return _from_source.name.rstrip(self.config_ext)
        else:
            return None

    def apply_special_attributes_configs(self, config):
        file_base = self.name
        file_extension = self.file_extension
        file_path = self.get_file_path(absolute=True)
        find_replace = {'_file_system_base':file_base, '_file_extension':file_extension,'_file_path':file_path}
        find_replace_values(config,find_replace)

    @property
    def config_inherited(self):
        if self.parent is None:
            return {}
        else:
            return self.parent.config_combined
    
    @property
    def config_combined(self):
        if self.config_explicit is None:
            config_combined = self.config_inherited
        else:
            config_combined = merge_dicts_by_top_level_keys(self.config_inherited,self.config_explicit)
        self.apply_special_attributes_configs(config_combined)
        return config_combined

class ConfigContainer(GenericNode, BranchNodeMixin, ConfigMixin):
    def __init__(self, parent=None, config_target=None, config_source=None):
        ConfigMixin.__init__(self, config_target,config_source)
        super().__init__(self.container_name, parent)
        BranchNodeMixin.__init__(self)

    def display_config(self):
        ConfigNode.display_config(self)
        for child in self.children.values():
            child.display_config()

    @property
    def file_extension(self):
        file_extension = None
        if len(self.children) > 0:
            first_child = next(iter(self.children.values()))
            if isinstance(first_child,ConfigNode):
                file_extension = first_child.file_extension
        return file_extension
    
class ConfigNode(GenericNode, LeafNodeMixin, ConfigMixin):
    def __init__(self, parent, config_target=None, config_source=None):
        ConfigMixin.__init__(self, config_target,config_source)
        super().__init__(self.container_name, parent)
        LeafNodeMixin.__init__(self)

    def display_config(self):
        print(self.get_file_path(), self.config_combined)

    @property
    def file_extension(self):
        node = self.parent
        file_extension =  node.name.split(".")[-1]
        return file_extension

class ConfigTree(GenericTree):
    def __init__(self, config_target, config_ext='.eel.yml'):
        self.config_ext= config_ext
        self.add_container(config_target=config_target)
        self.populate_tree(config_target,parent=self.root)

    def get_config_source(self,target):
        if isinstance(target, Folder):
            if self.config_ext in target.children:
                return target.children[self.config_ext].children['0']
        elif isinstance(target, File):
            if not target.name.endswith(self.config_ext):
                config_source_path = target.get_file_path() + self.config_ext
                if config_source_path in target.siblings:
                    return target.siblings[config_source_path].children['0']
            else:
                return target.children['0']
        elif isinstance(target, FilePart):
            parent = target.parent
            if not parent.name.endswith(self.config_ext):
                config_source_file_path = parent.get_file_path() + self.config_ext
                if config_source_file_path in parent.siblings:
                    config_source_file = parent.siblings[config_source_file_path]
                    for child in config_source_file.children.values():
                        if 'sub-path' in child.content:
                            if target.name == child.content['sub-path']:
                                return child
            else:
                return target
        return None
    
    def populate_tree(self,file_node,parent):
        skip_configs = []
        for child in file_node.children.values():
            if isinstance(child,Folder):
                fs_container = child
                eel_container=self.add_container(config_target=fs_container,parent=parent)
                self.populate_tree(child,parent=eel_container)
            elif isinstance(child,File):
                if not child.name.endswith(self.config_ext):
                    skip_configs.append(child.name + self.config_ext)
                    fs_container = child
                    eel_container=self.add_container(config_target=fs_container,parent=parent)
                    self.populate_tree(child,parent=eel_container)
                elif not (child.name in skip_configs or child.name == self.config_ext):
                    #TODO complete
                    print('sole ymls not covered')
                    print(child.name)
            elif isinstance(child,FilePart):
                fs_container = child
                eel_container=self.add_part(config_target=fs_container,parent=parent)
    
    def add_container(self, parent=None, config_target=None,config_source=None ):
        if not config_source:
            config_source = self.get_config_source(target= config_target)
        container = ConfigContainer(config_target=config_target, config_source=config_source, parent=parent)
        return self.add_node(container, parent)

    def add_part(self, parent=None, config_target=None,config_source=None ):
        if not config_source:
            config_source = self.get_config_source(target= config_target)
        node = ConfigNode(config_target=config_target, config_source=config_source, parent=parent)
        return self.add_node(node, parent)

    def display_configs(self):
        self.root.display_config()


# def populate_config(config_tree, basepath, parent=None):
#     pass

root_path = 'D:\\test_data'

# Example usage:
ft = FileTree(root_path)

# Replace 'your_directory_path' with the path to the directory you want to populate
populate_tree(ft, root_path)

# ct = ConfigTree(ft.root)
# ct.display_configs()

ft.display_tree()

# for i in ft.index.keys():
#     print(i)