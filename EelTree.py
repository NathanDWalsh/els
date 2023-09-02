import os
import zipfile
import yaml
from openpyxl import load_workbook
import logging
import re
import pandas as pd
from typing import Callable, Optional, Union

from ContentAwarePath import ContentAwarePath as CAPath

import EelIngest as ee
import EelConfig as ec
import EelFlow as ef


class GenericNode:
    def __init__(self, path: CAPath, parent: "GenericNode" = None, config: dict = {}):
        self.path = path
        self.parent = parent
        self.config_explicit = config

    @property
    def config_inherited(self) -> ec.Config:
        if self.parent is None:
            return ec.Config()
        else:
            return self.parent.config_combined

    @property
    def config_combined(self) -> ec.Config:
        config_inherited = self.config_inherited.model_copy(deep=True)
        if self.config_explicit is None:
            return config_inherited
        else:
            return merge_configs(config_inherited, self.config_explicit)
            # config_with_dicts = ec.deep_merge(config_inherited, self.config_explicit)
            # # config_as_dict = config_with_dicts.model_dump()
            # return ec.Config(config_with_dicts)
        # return self.config_inherited.model_copy(deep=True, update=self.config_explicit.model_dump(exclude_unset=True))

    @property
    def config(self) -> ec.Config:
        config = self.config_combined.model_copy(deep=True)
        config.sub_path = self.path.str
        # config.finalize()
        # print(config)
        # config.convert_special_attributes()
        # print(config)
        return config.eval_dynamic_attributes()

    @property
    def has_children(self) -> bool:
        if isinstance(self, BranchNodeMixin) and len(self.children) > 0:
            return True
        else:
            return False

    @property
    def siblings(self):
        return self.parent.children

    @property
    def base(self):
        if self.parent is None:
            return self
        else:
            return self.parent.base

    @property
    def all_base_leafs(self):
        return self.base.all_leafs

    @property
    def is_homog_branch(self):
        return (
            isinstance(self, BranchNodeMixin)
            and self.leaf_tables_same
            and self.leaf_count > 0
        )

    @property
    def is_non_empty_branch(self):
        return isinstance(self, BranchNodeMixin) and self.leaf_count > 0


class LeafNodeMixin:
    def __init__(self):
        # self.tree_part = TreePart.LEAF
        pass

    def display(self, item_path=None):
        file_path = self.path.str
        # if file_path == 'export_g2n_veeva_mx_w_prods.tsv\\0':
        if item_path:
            output = get_dict_item(self.config.model_dump(), item_path)
        else:
            output = self.config.model_dump()
        print(file_path, output)

    @property
    def type(self):
        return self.config.source.type

    @property
    def table(self):
        if self.config.target:
            return self.config.target.table
        else:
            return None

    @property
    def load_parallel(self):
        return self.config.source.load_parallel

    # def get_eel_executor(self, execute_fn=ee.ingest):
    #     item_res = self._get_tree_path()
    #     return ef.EelExecute(item_res, self.config, execute_fn)


class BranchNodeMixin:
    def __init__(self):
        # self.tree_part = TreePart.BRANCH
        self.children = {}

    def display(self, item_path=None):
        LeafNodeMixin.display(self, item_path)
        for child in self.children.values():
            child.display(item_path)

    def add_child(self, child_node):
        self.children[child_node.path.name] = child_node

    def remove_child(self, child_node):
        if child_node.path.name in self.children:
            del self.children[child_node.path.name]

    @property
    def all_leafs(self):
        for child in self.children.values():
            if isinstance(child, LeafNodeMixin):
                yield child
            else:
                yield from child.all_leafs

    @property
    def size(self):
        res = 0
        for leaf in self.all_leafs:
            res += leaf.size
        return res

    @property
    def size_weighted(self):
        res = 0
        for leaf in self.all_leafs:
            res += leaf.size * len(leaf.siblings)
        return res

    @property
    def all_file_count(self):
        res = 0
        for file in self.all_files:
            res += 1
        return res

    @property
    def all_files(self):
        if isinstance(self, File):
            yield self
        else:
            for child in self.children.values():
                if isinstance(child, File):
                    yield child
                else:
                    yield from child.all_files

    @property
    def leaf_count(self):
        count = 0
        for leaf in self.all_leafs:
            count += 1
        return count

    @property
    def leaf_tables_same(self):
        last_target = None
        for leaf in self.all_leafs:
            this_target = leaf.table
            if (not last_target is None) and last_target != this_target:
                return False
            last_target = leaf.table
        return True

    @property
    def leaf_parallels_same(self):
        last_target = None
        for leaf in self.all_leafs:
            this_target = leaf.load_parallel
            if (not last_target is None) and last_target != this_target:
                return False
            last_target = leaf.load_parallel
        return True

    @property
    def first_leaf(self):
        for child in self.children.values():
            if isinstance(child, LeafNodeMixin):
                return child
            else:
                return child.first_leaf

    @property
    def load_parallel(self):
        if self.config and self.config.source:
            return self.config.source.load_parallel
        else:
            return False

    @property
    def n_jobs(self):
        if self.load_parallel:
            cpu_cores = os.cpu_count()
            # proportion = self.size_weighted / self.root.size_weighted
            # res = min(max(round(proportion * cpu_cores),1),self.leaf_count)

            res = min(self.leaf_count, cpu_cores)
            return res
        else:
            return 1

    @property
    def children_size_asc(self):
        # Yield the current node's children sorted by size
        for child in sorted(self.children.values(), key=lambda x: x.size, reverse=True):
            yield child

    # def get_eel_executors(self, execute_fn=ei.ingest):
    #     eel_executers = []
    #     for leaf in self.all_leafs:
    #         eel_exec = leaf.get_eel_executor(execute_fn=ei.ingest)
    #         eel_executers.append(eel_exec)
    #     return eel_executers

    @property
    def get_leaf_df(self) -> pd.DataFrame:
        def leaf_to_dict(leaf):
            data = {}
            data["name"] = leaf.path.name
            data["file_path"] = leaf.path.file.abs.str
            data["type"] = leaf.type
            data["table"] = leaf.table
            data["load_parallel"] = leaf.load_parallel
            data["config"] = leaf.config

            return data

        data = [leaf_to_dict(leaf) for leaf in self.all_leafs]
        df = pd.DataFrame(data)
        return df

    def get_file_wrappers(
        self, df: pd.DataFrame, execute_fn: Callable[[ec.Config], bool]
    ) -> list[ef.EelFileWrapper]:
        res = []
        for file, file_gb in df.groupby(["file_path", "type"]):
            executes = []
            if file[1] == ".xlsx":
                file_wrapper = ef.EelXlsxWrapper(file[0], ef.EelFlow(executes, 1))
            else:
                file_wrapper = ef.EelFileWrapper(file[0], ef.EelFlow(executes, 1))
            for task_row in file_gb[["name", "config"]].itertuples():
                task_flow = ef.EelExecute(task_row.name, task_row.config, execute_fn)
                executes.append(task_flow)
            res.append(file_wrapper)
        return res

    def get_ingest_taskflow(self) -> ef.EelFlow:
        df = self.get_leaf_df
        root_flows = []
        res = ef.EelFlow(root_flows, 1)
        for _, table_gb in df.groupby("table", dropna=False):
            file_group_flows = self.get_file_wrappers(table_gb, ee.ingest)
            file_group_wrapper = ef.EelFileGroupWrapper(file_group_flows, False)
            root_flows.append(file_group_wrapper)
        return res

    def get_detect_taskflow(self) -> ef.EelFlow:
        df = self.get_leaf_df
        root_flows = self.get_file_wrappers(df, ee.detect)
        res = ef.EelFlow(root_flows, 1)
        return res


class Folder(GenericNode, BranchNodeMixin):
    def __init__(self, path: str, parent: BranchNodeMixin, config: ec.Config):
        super().__init__(path, parent, config)
        BranchNodeMixin.__init__(self)


class File(GenericNode, BranchNodeMixin):
    def __init__(self, path: str, parent: Folder, config: ec.Config):
        super().__init__(path, parent, config)
        BranchNodeMixin.__init__(self)


class FilePart(GenericNode, LeafNodeMixin):
    def __init__(self, path: str, parent: File, config: ec.Config):
        super().__init__(path, parent, config)
        LeafNodeMixin.__init__(self)
        self.size = None
        # self.content = content


class EelTree:
    CONFIG_FILE_EXT = ".eel.yml"
    FOLDER_CONFIG_FILE_STEM = "_"
    CONFIG_PREVIEW_FILE_NAME = "_preview.eel.yml"

    def __init__(self, path: CAPath) -> None:
        self.populate_tree(path)

    def populate_tree(self, path: CAPath, parent: BranchNodeMixin = None) -> None:
        if path.is_dir():
            config = self.get_paired_config(path)
            folder = self.add_folder(path, parent, config)
            ignore_configs = [
                path / self.folder_config_name,
                CAPath() / self.CONFIG_PREVIEW_FILE_NAME,
            ]
            for path_item in path.iterdir():
                if not path_item.str.endswith(self.CONFIG_FILE_EXT):
                    ignore_configs.append(CAPath(path_item.str + self.CONFIG_FILE_EXT))
                    self.populate_tree(path_item, parent=folder)
                elif not path_item in ignore_configs:
                    # TODO complete
                    logging.error("ERROR: sole ymls not covered: " + path_item.str)
        elif path.str.find("~$") == -1:  # exclude hidden excel temps, TODO not ideal
            config = self.get_paired_config(path)
            file = self.add_file(path, parent, config)
            if path.suffix == ".xlsx":
                sheet_deets = get_sheeet_deets(path.str)
                for ws_name, ws_size in sheet_deets.items():
                    config = self.get_paired_config(path, ws_name)
                    self.add_file_part(file.path / ws_name, file, config, ws_size)
            else:
                file_size = path.stat().st_size
                self.add_file_part(file.path / file.path.stem, file, size=file_size)

    @property
    def folder_config_name(self):
        return self.FOLDER_CONFIG_FILE_STEM + self.CONFIG_FILE_EXT

    def add_base_node(self, node: GenericNode):
        self.base = node
        self.index = {}
        self.index["."] = self.base
        return node

    def add_node(self, node: GenericNode, parent: GenericNode) -> GenericNode:
        if parent is None:
            return self.add_base_node(node)
        else:
            parent.add_child(node)  # TODO, move to GenericNode init
            self.add_index(node)
            return node

    def add_index(self, node: GenericNode):
        path = node.path.str
        if path in self.index:
            print(
                "Warning, duplicate data path found, index will reference most recent: "
                + path.str
            )
        self.index[path] = node

    def remove_node(self, node):
        if node.parent_folder:
            node.parent_folder.remove_child(node)
        else:
            self.base.remove_child(node)
        del self.index[node.path.str]

    def get_node_by_path(self, file_path):
        # relative_file_path = file_path.replace(os.path.join(self.root.path, ""), "")
        return self.index.get(file_path)

    def display_tree(self, item_path=None):
        self.base.display(item_path)

    def save_eel_yml_preview(self):
        ymls = []
        for path, node in self.index.items():
            node_config = node.config.model_dump(exclude_none=True)
            node_config["sub_path"] = path
            if node.parent is None:
                save_yml_dict = node_config
            else:
                parent_config = node.parent.config.model_dump(exclude_none=True)
                save_yml_dict = dict_diff(parent_config, node_config)
            # if "source" in save_yml_dict:
            #     if "type" in save_yml_dict["source"]:
            #         del save_yml_dict["source"]["type"]
            #     if "file_path" in save_yml_dict["source"]:
            #         del save_yml_dict["source"]["file_path"]
            #     if len(save_yml_dict["source"]) == 0:
            #         del save_yml_dict["source"]
            # if "file_system_base" in save_yml_dict:
            #     del save_yml_dict["file_system_base"]
            # if "file_extension" in save_yml_dict:
            #     del save_yml_dict["file_extension"]
            # if "file_path" in save_yml_dict:
            #     del save_yml_dict["file_path"]
            ymls.append(save_yml_dict)
        save_path = self.base.path / self.CONFIG_PREVIEW_FILE_NAME
        with save_path.open("w", encoding="utf-8") as file:
            yaml.safe_dump_all(ymls, file, sort_keys=False, allow_unicode=True)

    # def get_paired_config(self, item_path: str, sub_path: str = ".") -> ec.Config:
    def get_paired_config(self, item_path: CAPath, sub_path: str = ".") -> dict:
        config_path = self.get_paired_config_path(item_path)
        if config_path:
            ymls = get_yml_docs(config_path)
            configs = get_configs(ymls)
            for config, yml in zip(configs, ymls):
                if sub_path == config.sub_path:
                    return yml
            return {}
            # return ec.Config()

    def get_paired_config_path(self, path: CAPath) -> Optional[CAPath]:
        if path.is_dir():
            config_path = path / self.folder_config_name
        elif path.is_file():
            if path.str.endswith(self.CONFIG_FILE_EXT):
                return path
            else:
                config_path = CAPath(path.str + self.CONFIG_FILE_EXT)
        if config_path.exists():
            return config_path
        return None

    def add_folder(self, path: CAPath, parent, config: dict) -> Folder:
        # if not parent is None:
        #     path = os.path.basename(path)
        folder = Folder(path, parent, config)
        return self.add_node(folder, parent)

    def add_file(self, path: CAPath, parent, config: dict) -> File:
        # if not parent is None:
        #     path = os.path.basename(path)
        file = File(path, parent, config)
        return self.add_node(file, parent)

    def add_file_part(
        self, path: CAPath, parent, config: dict = {}, size: int = None
    ) -> FilePart:
        file_part = FilePart(path, parent, config)
        file_part.size = size
        return self.add_node(file_part, parent)


def dict_diff(dict1: dict, dict2: dict) -> dict:
    """
    Return elements that are in dict2 but not in dict1.

    :param dict1: First dictionary
    :param dict2: Second dictionary
    :return: A dictionary with elements only from dict2 that are not in dict1
    """
    diff = {}

    for key, value in dict2.items():
        # If key is not present in dict1, add the item
        if key not in dict1:
            diff[key] = value
        # If key is present in both dictionaries and both values are dictionaries, recurse
        elif isinstance(value, dict) and isinstance(dict1[key], dict):
            nested_diff = dict_diff(dict1[key], value)
            if nested_diff:
                diff[key] = nested_diff
        # If the key's value is different between the dictionaries, add the item
        elif dict1[key] != value:
            diff[key] = value

    return diff


def get_dict_item(dict_: dict, item_path: list, default=None):
    current_key = item_path[0]
    if current_key in dict_:
        next_dict = dict_[current_key]
        remaining_path = item_path[1:]
        if remaining_path:
            return get_dict_item(next_dict, remaining_path, default)
        else:
            return next_dict
    else:
        return default


def merge_configs(*configs: list[Union[ec.Config, dict]]) -> ec.Config:
    dicts = []
    for config in configs:
        if isinstance(config, ec.Config):
            dicts.append(config.model_dump())
        elif isinstance(config, dict):
            dicts.append(config)
        else:
            raise Exception(
                "merge_configs: configs should be a list of Configs or dicts"
            )
    dict_result = merge_dicts_by_top_level_keys(*dicts)
    res = ec.Config(**dict_result)
    return res


def merge_dicts_by_top_level_keys(*dicts: list[dict]) -> dict:
    merged_dict = {}
    for dict_ in dicts:
        for key, value in dict_.items():
            if (
                key in merged_dict
                and isinstance(value, dict)
                and (not merged_dict[key] is None)
            ):
                merged_dict[key].update(value)
            elif not value is None:
                # Add a new key-value pair to the merged dictionary
                merged_dict[key] = value
    return merged_dict


def natural_sort_key(s):
    return [
        int(text) if text.isdigit() else text.lower() for text in re.split(r"(\d+)", s)
    ]


def get_sheet_sizes(file_path):
    sheet_sizes = []

    with zipfile.ZipFile(file_path, "r") as zip_file:
        sheet_names = zip_file.namelist()

        for sheet_name in sorted(sheet_names, key=natural_sort_key):
            if sheet_name.startswith("xl/worksheets/sheet"):
                file_info = zip_file.getinfo(sheet_name)
                sheet_sizes.append(file_info.file_size)

    return sheet_sizes


def get_sheet_names(file_path):
    workbook = load_workbook(
        filename=file_path, read_only=True, data_only=True, keep_links=False
    )
    worksheet_names = workbook.sheetnames
    workbook.close()
    return worksheet_names


def get_sheeet_deets(file_path):
    # TODO: visible_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'visible']
    names = get_sheet_names(file_path)
    sizes = get_sheet_sizes(file_path)
    return dict(zip(names, sizes))


def get_yml_docs(path: CAPath) -> list[dict]:
    with path.open() as file:
        yaml_text = file.read()
        documents = list(yaml.safe_load_all(yaml_text))
    return documents


def get_configs(ymls: list[dict]) -> list[ec.Config]:
    configs = []
    for yml in ymls:
        config = ec.Config(**yml)
        configs.append(config)
    return configs


# def get_yml_configs(file_path: str) -> list[ec.Config]:
#     ymls = get_yml_docs(file_path)
#     configs = get_configs(ymls)
#     return configs

import warnings

warnings.simplefilter("error")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(relativeCreated)d - %(message)s")
    logging.info("Getting Started")

    os.chdir("D:\\test_data4")
    base_path = CAPath()

    ft = EelTree(base_path)
    ft.display_tree()

    # print(type(ec.Config(**ft.base.first_leaf.config.model_dump())))
    # print(ec.Config(**ft.base.first_leaf.config.model_dump()))

    # print(ft.base.first_leaf.config)
    # print(ft.base.config.eval_dynamic_attributes())

    logging.info("Tree Created")
    # ft.save_eel_yml_preview()

    taskflow = ft.base.get_ingest_taskflow()
    print(taskflow.to_tuple)
    taskflow.execute()

    logging.info("Fin")
