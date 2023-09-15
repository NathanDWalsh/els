from pathlib import Path
from anytree import NodeMixin, RenderTree, PreOrderIter
import pandas as pd
import os
from typing import Union, Callable, List, Optional
from openpyxl import load_workbook
import yaml
import logging

import eel.config as ec
import eel.flow as ef
import eel.execute as ee
from eel.pathprops import HumanPathPropertiesMixin

CONFIG_FILE_EXT = ".eel.yml"
FOLDER_CONFIG_FILE_STEM = "_"
ROOT_CONFIG_FILE_STEM = "__"


def get_folder_config_name():
    return FOLDER_CONFIG_FILE_STEM + CONFIG_FILE_EXT


def get_root_config_name():
    return ROOT_CONFIG_FILE_STEM + CONFIG_FILE_EXT


class PathToStringMixin:
    @property
    def str(self):
        return str(self)


class ConfigInheritanceMixin:
    @property
    def config(self) -> ec.Config:
        config_line = []
        # if root eel config is mandatory, this "default dump line" is not required
        config_line.append(ec.Config().model_dump(exclude_none=True))

        for node in self.ancestors + (self,):
            if node._config:
                config_line.append(node._config)
        config_merged = ConfigInheritanceMixin.merge_configs(*config_line)
        config_copied = config_merged.model_copy(deep=True)
        # config_copied.sub_path = self.str
        config_evaled = self.eval_dynamic_attributes(config_copied)
        return config_evaled

    def get_path_props_find_replace(self) -> dict:
        res = {}
        for member in ec.DynamicPathValue:
            path_val = getattr(self, member.value[1:])
            res[member.value] = path_val
        return res

    def eval_dynamic_attributes(self, config: ec.Config) -> ec.Config:
        config_dict = config.model_dump(exclude_none=True)
        find_replace = self.get_path_props_find_replace()
        ContentAwarePath.swap_dict_vals(config_dict, find_replace)
        res = ec.Config(**config_dict)
        return res

    @staticmethod
    def swap_dict_vals(dictionary: dict, find_replace_dict: dict) -> None:
        for key, value in dictionary.items():
            if isinstance(value, dict):
                ContentAwarePath.swap_dict_vals(dictionary[key], find_replace_dict)
            elif isinstance(value, list):
                pass
            elif value in find_replace_dict:
                dictionary[key] = find_replace_dict[value]

    @staticmethod
    def merge_configs(*configs: list[Union[ec.Config, dict]]) -> ec.Config:
        dicts = []
        for config in configs:
            if isinstance(config, ec.Config):
                dicts.append(config.model_dump())
            elif isinstance(config, dict):
                dicts.append(config)
            else:
                raise Exception("configs should be a list of Configs or dicts")
        dict_result = ConfigInheritanceMixin.merge_dicts_by_top_level_keys(*dicts)
        res = ec.Config(**dict_result)
        return res

    @staticmethod
    def merge_dicts_by_top_level_keys(*dicts: list[dict]) -> dict:
        merged_dict = {}
        for dict_ in dicts:
            for key, value in dict_.items():
                if (
                    key in merged_dict
                    and isinstance(value, dict)
                    and (not merged_dict[key] is None)
                ):
                    merged_dict[key].update(value)
                elif value is not None:
                    # Add a new key-value pair to the merged dictionary
                    merged_dict[key] = value
        return merged_dict


class ContentAwarePath(
    Path,
    PathToStringMixin,
    HumanPathPropertiesMixin,
    NodeMixin,
    ConfigInheritanceMixin,
):
    _flavour = type(Path())._flavour

    def __init__(self, *args, parent=None, **kwargs):
        self.parent = parent
        self._config = self.get_paired_config()

    def get_paired_config(self) -> dict:
        if (
            self.parent
            and self.parent._config
            and ("children" in self.parent._config)
            and isinstance(self.parent._config["children"], list)
            and isinstance(self.parent._config["children"][0], dict)
        ):
            if self.str in self.parent._config["children"]:
                return self.parent._config["children"][self.str]
            else:
                logging.error("Config not found in parent config when expected")
                return {}
        elif not self.is_content():
            config_path = self.get_paired_config_path()
            if config_path:
                ymls = get_yml_docs(config_path)
                # configs are loaded only to ensure they conform with yml schema
                configs = get_configs(ymls)
                if len(configs) > 1:
                    logging.error(
                        "Found more than one yml document, using first one only"
                    )
                yml = ymls[0]
                return yml
        else:
            return {}

    def get_paired_config_path(self) -> Optional[Path]:
        if self.is_root:
            config_path = self / get_root_config_name()
        elif self.is_dir():
            config_path = self / get_folder_config_name()
        elif self.is_file():
            if self.str.endswith(CONFIG_FILE_EXT):
                return self
            else:
                config_path = Path(self.str + CONFIG_FILE_EXT)
        if config_path.exists():
            return config_path
        return None

    def display_tree(self):
        for pre, fill, node in RenderTree(self):
            print("%s%s" % (pre, node.name))

    @property
    def parent(self):
        # return NodeMixin().parent
        return NodeMixin.parent.fget(self)

    @property
    def root(self):
        return NodeMixin.root.fget(self)

    @parent.setter
    def parent(self, value):
        NodeMixin.parent.fset(self, value)

    # @property
    def is_content(self) -> bool:
        """
        Check if the path points to a content inside a file.
        A naive check is to see if the parent exists as a file.
        """
        if self.is_root:
            return False
        else:
            return self.parent.is_file()

    @property
    def subdir_patterns(self) -> List[str]:
        # TODO patterns may overlap
        res = (
            self._config["children"]
            if self._config and "children" in self._config
            else None
        )
        if res is None:
            res = ["*"]
        elif isinstance(res, str):
            res = [res]
        # if list of dicts
        elif isinstance(res, list) and all(isinstance(item, dict) for item in res):
            # get key of each dict as list entries
            res = [next(iter(d)) for d in res]
        return res

    def count_recursive_files_in_subdirs(self) -> int:
        count = 0
        # , subdir_patterns: Union[List[str], None, str, dict]

        for pattern in self.subdir_patterns:
            for subpath in self.glob(pattern):
                if subpath.is_dir():
                    count += sum(
                        1
                        for file in subpath.rglob("*")
                        if file.is_file() and not file.is_hidden()
                    )
                elif subpath.is_file() and not subpath.is_hidden():
                    count += 1
        return count

    def iterbranch(self) -> "ContentAwarePath":
        if self.is_dir():
            skip_paths = []
            skip_paths.append(self / get_folder_config_name())
            skip_paths.append(self / get_root_config_name())
            # dirs allow glob matches
            for pattern in self.subdir_patterns:
                for subpath in self.glob(pattern):
                    subpath._config = subpath.get_paired_config()
                    if (subpath not in skip_paths) and (
                        (subpath.is_file() and not subpath.is_hidden())
                        or (subpath.count_recursive_files_in_subdirs())
                    ):  # not subpath.str.endswith(CONFIG_FILE_EXT) and
                        # paths will not repeat due to overlapping globs
                        skip_paths.append(subpath)
                        # do not yield a paired yml
                        # TODO: assumes pathlib sorts ascending
                        skip_paths.append(
                            ContentAwarePath(subpath.str + CONFIG_FILE_EXT)
                        )
                        yield subpath
        elif self.is_file():
            # content allows exact matches
            leaf_names = self.get_content_leaf_names()

            for pattern in self.subdir_patterns:
                if pattern == "*":
                    for leaf in leaf_names:
                        yield ContentAwarePath(self / leaf, parent=self)
                    break
                if pattern in leaf_names:
                    yield ContentAwarePath(self / pattern, parent=self)

    def get_content_leaf_names(self) -> List[str]:
        if self.suffix == ".xlsx":
            return get_sheet_names(self.str)
        elif self.suffix == ".yml":  # and self._config.type =='mssql'
            # return get_db_tables
            pass  # TODO
        else:
            return [self.stem]

    def is_hidden(path: Path) -> bool:
        """Check if the given Path object is hidden."""
        # Check for UNIX-like hidden files/directories
        if path.name.startswith("."):
            return True

        # Check for Windows hidden files/directories
        if os.name == "nt":
            try:
                attrs = os.stat(path)
                return attrs.st_file_attributes & os.FILE_ATTRIBUTE_HIDDEN
            except AttributeError:
                # If FILE_ATTRIBUTE_HIDDEN not defined,
                # assume it's not hidden
                pass

        return False

    @property
    def abs(self) -> "ContentAwarePath":
        return ContentAwarePath(self.absolute())

    @property  # fs = filesystem, can return a File or Dir but not content
    def fs(self) -> "ContentAwarePath":
        if self.is_content():
            res = self.parent
        else:
            res = self
        return res

    @property
    def dir(self) -> "ContentAwarePath":
        if self.is_content():
            res = self.parent.parent
        elif self.is_file():
            res = self.parent
        else:
            res = self
        return res

    @property
    def file(self) -> "ContentAwarePath":
        if self.is_content():
            res = self.parent
        elif self.is_file():
            res = self
        else:
            res = None
        return res

    @property
    def ext(self) -> str:
        file = self.file
        if file:
            return file.suffix
        else:
            return ""

    @property
    def get_leaf_df(self) -> pd.DataFrame:
        def leaf_to_dict(leaf):
            data = {}
            data["name"] = leaf.name
            data["file_path"] = leaf.file.abs.str
            data["type"] = leaf.config.source.type
            data["table"] = leaf.config.target.table
            data["load_parallel"] = leaf.config.source.load_parallel
            data["config"] = leaf.config

            return data

        data = [leaf_to_dict(leaf) for leaf in self.leaves]
        df = pd.DataFrame(data)
        return df

    @staticmethod
    def apply_file_wrappers(
        parent: ef.FlowNodeMixin,
        df: pd.DataFrame,
        execute_fn: Callable[[ec.Config], bool],
    ) -> None:
        ingest_files = ef.EelFlow(parent=parent, n_jobs=1)
        for file, file_gb in df.groupby(["file_path", "type"]):
            if file[1] == ".xlsx":
                file_wrapper = ef.EelXlsxWrapper(parent=ingest_files, file_path=file[0])
            else:
                file_wrapper = ef.EelFileWrapper(parent=ingest_files, file_path=file[0])
            exe_flow = ef.EelFlow(parent=file_wrapper, n_jobs=1)
            for task_row in file_gb[["name", "config"]].itertuples():
                ef.EelExecute(
                    parent=exe_flow,
                    name=task_row.name,
                    config=task_row.config,
                    execute_fn=execute_fn,
                )

    def get_ingest_taskflow(self) -> ef.EelFlow:
        df = self.get_leaf_df
        root_flow = ef.EelFlow()
        for table, table_gb in df.groupby("table", dropna=False):
            file_group_wrapper = ef.EelFileGroupWrapper(
                parent=root_flow, name=table, exec_parallel=False
            )
            ContentAwarePath.apply_file_wrappers(
                parent=file_group_wrapper, df=table_gb, execute_fn=ee.ingest
            )

        return root_flow

    def get_detect_taskflow(self) -> ef.EelFlow:
        df = self.get_leaf_df
        root_flows = ContentAwarePath.apply_file_wrappers(df, ee.detect)
        res = ef.EelFlow(root_flows, 1)
        return res

    def get_eel_yml_preview(self, diff: bool = True) -> List[dict]:
        ymls = []
        # for path, node in self.index.items():
        for node in [node for node in PreOrderIter(self)]:
            node_config = node.config.model_dump(exclude_none=True)
            # node_config["sub_path"] = node.str
            if node.is_root:
                save_yml_dict = node_config
            else:
                parent_config = node.parent.config.model_dump(exclude_none=True)
                if diff:
                    save_yml_dict = dict_diff(parent_config, node_config)
                else:
                    save_yml_dict = node_config
            ymls.append(save_yml_dict)
        return ymls
        # save_path = self.root.path / self.CONFIG_PREVIEW_FILE_NAME
        # with save_path.open("w", encoding="utf-8") as file:
        #     yaml.safe_dump_all(ymls, file, sort_keys=False, allow_unicode=True)


def dict_diff(dict1: dict, dict2: dict) -> dict:
    """
    Return elements that are in dict2 but not in dict1.

    :param dict1: First dictionary
    :param dict2: Second dictionary
    :return: A dictionary with elements only from dict2 that are not in dict1
    """
    diff = {}

    for key, value in dict2.items():
        # If key is not present in dict1, add the item
        if key not in dict1:
            diff[key] = value
        # If key is present in both dicts and both values are dicts, recurse
        elif isinstance(value, dict) and isinstance(dict1[key], dict):
            nested_diff = dict_diff(dict1[key], value)
            if nested_diff:
                diff[key] = nested_diff
        elif dict1[key] != value:
            diff[key] = value

    return diff


def get_sheet_names(file_path, sheet_states: list = ["visible"]) -> List[str]:
    workbook = load_workbook(
        filename=file_path, read_only=True, data_only=True, keep_links=False
    )

    if sheet_states is None:
        worksheet_names = workbook.sheetnames
    else:
        worksheet_names = [
            sheet.title
            for sheet in workbook.worksheets
            if sheet.sheet_state in sheet_states
        ]

    workbook.close()
    return worksheet_names


def get_yml_docs(path: ContentAwarePath) -> list[dict]:
    with path.open() as file:
        yaml_text = file.read()
        documents = list(yaml.safe_load_all(yaml_text))
    return documents


def get_configs(ymls: list[dict]) -> list[ec.Config]:
    configs = []
    for yml in ymls:
        config = ec.Config(**yml)
        configs.append(config)
    return configs


def grow_branches(
    path: ContentAwarePath = ContentAwarePath(), parent: ContentAwarePath = None
) -> ContentAwarePath:
    if path.is_dir() or path.is_file():
        node = ContentAwarePath(path, parent=parent)
        for path_item in node.iterbranch():
            grow_branches(path_item, parent=node)
        return node
    # TODO: consider how database connections / ymls will be handled
