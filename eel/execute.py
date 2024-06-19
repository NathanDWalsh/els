import pandas as pd
import sqlalchemy as sa
import logging
from typing import Union, Optional
import eel.config as ec
import os
import csv
from python_calamine import CalamineWorkbook
from openpyxl import load_workbook


open_files = {}
staged_frames = {}


def push_frame(df: pd.DataFrame, target: ec.Target, add_cols: dict) -> bool:
    res = False
    if df is not None:
        if not target or not target.type:
            logging.info("no target defined, printing first 100 rows:")
            print(df.head(100))
            res = True
        else:
            if target.type in (".csv"):
                res = push_csv(df, target, add_cols)
            if target.type in (".xlsx"):
                res = push_excel(df, target, add_cols)
            elif target.type in ("mssql", "postgres", "duckdb"):
                res = push_sql(df, target, add_cols)
            elif target.type in ("pandas"):
                staged_frames[target.table] = df
                res = True
            else:
                pass
    return res


def push_sql(source_df: pd.DataFrame, target: ec.Target, add_cols: dict) -> bool:
    if not target.db_connection_string:
        raise Exception("invalid db_connection_string")
    if not target.table:
        raise Exception("invalid to_sql")
    with sa.create_engine(
        target.db_connection_string, fast_executemany=True
    ).connect() as sqeng:
        if target.to_sql:
            kwargs = target.to_sql.model_dump()
        else:
            kwargs = {}
        source_df.to_sql(
            target.table,
            sqeng,
            target.dbschema,
            index=False,
            if_exists="append",
            **kwargs,
        )
        sqeng.connection.commit()
        return True


def push_csv(source_df: pd.DataFrame, target: ec.Target, add_cols: dict) -> bool:
    if not target.file_path_dynamic:
        raise Exception("no file path")
    if not os.path.exists(os.path.isfile(target.file_path_dynamic)):
        raise Exception("invalid file path")

    if target.to_csv:
        kwargs = target.to_csv.model_dump()
    else:
        kwargs = {}

    source_df.to_csv(
        target.file_path_dynamic, index=False, mode="a", header=False, **kwargs
    )

    return True


def push_excel(source_df: pd.DataFrame, target: ec.Target, add_cols: dict) -> bool:
    if not target.file_path_dynamic:
        raise Exception("invalid file_path")
    sheet_name = target.table or "Sheet1"

    if not sheet_name:
        raise Exception("sheet name not defined")

    # if target.to_excel:
    #     kwargs = target.to_excel.model_dump()
    # else:
    #     kwargs = {}

    sheet_height = get_excel_sheet_height(target.file_path_dynamic, sheet_name)
    start_row = sheet_height + 1

    if sheet_height is None:
        raise Exception("sheet name not found")

    with pd.ExcelWriter(
        target.file_path_dynamic, mode="a", if_sheet_exists="overlay"
    ) as writer:
        source_df.to_excel(
            writer, index=False, sheet_name=sheet_name, startrow=start_row, header=False
        )

    return True


def pull_sql(frame: ec.Frame, nrows=None):
    if not frame.db_connection_string:
        raise Exception("invalid db_connection_string")
    if not frame.sqn:
        raise Exception("invalid sqn")
    with sa.create_engine(frame.db_connection_string).connect() as sqeng:
        stmt = sa.select(sa.text("*")).select_from(sa.text(frame.sqn)).limit(nrows)
        df = pd.read_sql(stmt, con=sqeng)
    return df


def build_sql(df: pd.DataFrame, target: ec.Frame, add_cols: dict) -> bool:
    if not target.db_connection_string:
        raise Exception("invalid db_connection_string")
    if not target.sqn:
        raise Exception("invalid sqn")
    if not target.table:
        raise Exception("invalid table")

    with sa.create_engine(target.db_connection_string).connect() as sqeng:
        sqeng.execute(sa.text(f"drop table if exists {target.sqn}"))

        # Use the first row to create the table structure
        df.head(1).to_sql(target.table, sqeng, schema=target.dbschema, index=False)

        # Delete the temporary row from the table
        sqeng.execute(sa.text(f"DELETE FROM {target.sqn}"))

        if add_cols:
            for col_name, col_val in add_cols.items():
                if col_val == ec.DynamicColumnValue.ROW_INDEX.value:
                    # Add an identity column to the table
                    sqeng.execute(
                        sa.text(
                            (
                                f"ALTER TABLE {target.sqn} ADD {col_name}"
                                " int identity(1,1) PRIMARY KEY "
                            )
                        )
                    )

        sqeng.connection.commit()
    return True


def build_csv(df: pd.DataFrame, target: ec.Frame) -> bool:
    if not target.file_path_dynamic:
        raise Exception("invalid file_path")

    # save header row to csv, overwriting if exists
    df.head(0).to_csv(target.file_path_dynamic, index=False, mode="w")

    return True


def get_excel_sheet_height(file_path: str, sheet_name: str) -> int:
    workbook = CalamineWorkbook.from_path(file_path)
    if sheet_name in workbook.sheet_names:
        return workbook.get_sheet_by_name(sheet_name).total_height
    else:
        return None


def get_excel_sheet_row(file_path: str, sheet_name: str, row_index: int) -> list:
    workbook = CalamineWorkbook.from_path(file_path)
    if sheet_name in workbook.sheet_names:
        return workbook.get_sheet_by_name(sheet_name).to_python(nrows=row_index + 1)[-1]
    else:
        return None


def build_excel_frame(df: pd.DataFrame, target: ec.Frame) -> bool:
    if not target.file_path_dynamic:
        raise Exception("invalid file_path")
    sheet_name = target.table or "Sheet1"

    if not sheet_name:
        raise Exception("sheet name not defined")

    # save header row to csv, overwriting if exists
    # df.head(0).to_excel(target.file_path_dynamic, index=False, mode="w")

    kwargs = {}

    if not os.path.exists(target.file_path_dynamic):
        kwargs["mode"] = "w"
    else:
        kwargs["mode"] = "a"
        kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(target.file_path_dynamic, **kwargs) as writer:
        df.head(0).to_excel(writer, index=False, sheet_name=sheet_name)

    return True


# def build_excel_table(df: pd.DataFrame, target: ec.Frame, add_cols: dict) -> bool:
#     if not target.file_path:
#         raise Exception("invalid file_path")
#     if not target.table:
#         raise Exception("invalid table")

#     # save header row to csv, overwriting if exists
#     df.head(0).to_csv(target.file_path, index=False, mode="w")

#     return True


def build_target(df: pd.DataFrame, target: ec.Frame, add_cols: dict) -> bool:
    if target.type in ("mssql", "postgres", "duckdb"):
        res = build_sql(df, target, add_cols)
    elif target.type in (".csv"):
        res = build_csv(df, target)
    elif target.type in (".xlsx"):
        # res = df.to_excel(target.file_path, index=False)
        res = build_excel_frame(df, target)
    elif target.type in ("pandas"):
        res = df
    else:
        raise Exception("invalid target type")
    return res


def truncate_target(target: ec.Target) -> bool:
    if target.type in ("mssql", "postgres", "duckdb"):
        res = truncate_sql(target)
    elif target.type in (".csv"):
        res = truncate_csv(target)
    elif target.type in (".xlsx"):
        res = truncate_excel(target)
    # elif target.type in ("pandas"):
    #     res = truncate_pandas_table(target)
    else:
        raise Exception("invalid target type")
    return res


def truncate_csv(target: ec.Target) -> bool:
    if not target.file_path_dynamic:
        raise Exception("no file path")
    if not os.path.exists(os.path.isfile(target.file_path_dynamic)):
        raise Exception("invalid file path")

    # read the first row of the file
    with open(target.file_path_dynamic, "r") as f:
        reader = csv.reader(f)
        first_row = next(reader)

    # write the first row back to the file
    with open(target.file_path_dynamic, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(first_row)

    return True


def truncate_excel(target: ec.Target) -> bool:
    if not target.file_path_dynamic:
        raise Exception("no file path")
    if not os.path.exists(os.path.isfile(target.file_path_dynamic)):
        raise Exception("invalid file path")

    # read the first row of the file
    with open(target.file_path_dynamic, "r") as f:
        reader = csv.reader(f)
        first_row = next(reader)

    # write the first row back to the file
    with open(target.file_path_dynamic, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(first_row)

    return True


def clear_excel_sheet_after_row(file_path, sheet_name, row_start):
    # Load the workbook and select the sheet
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    # Iterate over the rows
    for i in range(row_start, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            # Clear the cell
            sheet.cell(row=i, column=j).value = None

    # Save the workbook
    wb.save(file_path)


def truncate_sql(target: ec.Target) -> bool:
    if not target.db_connection_string:
        raise Exception("invalid db_connection_string")
    with sa.create_engine(target.db_connection_string).connect() as sqeng:
        sqeng.execute(sa.text(f"truncate table {target.sqn}"))
        sqeng.connection.commit()
    return True


def frames_consistent(config: ec.Config) -> bool:
    target, source, add_cols = get_configs(config)

    if target and target.type in ("pandas"):
        return True

    ignore_cols = []
    if add_cols:
        for k, v in add_cols.items():
            if v == ec.DynamicColumnValue.ROW_INDEX.value:
                ignore_cols.append(k)

    source_df = pull_frame(source, 100)
    source_df = add_columns(source_df, add_cols)
    target_df = pull_frame(target, 100)
    return data_frames_consistent(source_df, target_df, ignore_cols)


def data_frames_consistent(
    df1: pd.DataFrame, df2: pd.DataFrame, ignore_cols: list = []
) -> bool:
    res = True
    ignore_cols_set = set(ignore_cols)
    # Compare the column names and types
    source_cols = set(df1.columns.tolist()) - ignore_cols_set
    target_cols = set(df2.columns.tolist()) - ignore_cols_set

    if source_cols != target_cols:
        in_source = source_cols - target_cols
        in_target = target_cols - source_cols
        if in_source:
            logging.info("source has more columns:" + str(in_source))
        if in_target:
            logging.info("target has more columns:" + str(in_target))
        res = False
    else:
        for col in source_cols:
            # if nulls are returned from sql and object type is set in df
            if df2[col].dtype != "object" and df1[col].dtype != df2[col].dtype:
                logging.info(
                    f"{col} has a different data type source "
                    f"{df1[col].dtype} target {df2[col].dtype}"
                )
                res = False

    return res  # Table exists and has the same field names and types


def get_sql_data_type(dtype):
    if dtype == "int64":
        return "INT"
    elif dtype == "float64":
        return "FLOAT"
    elif dtype == "bool":
        return "BIT"
    elif dtype == "object":
        return "VARCHAR(MAX)"
    elif dtype == "datetime64":
        return "DATETIME"
    else:
        return "VARCHAR(MAX)"


def pull_csv(file_path, **kwargs):
    df = pd.read_csv(file_path, **kwargs)
    return df


def pull_excel(file_path, **kwargs):
    df = pd.read_excel(file_path, **kwargs)
    return df


def get_source_kwargs(
    read_x, frame: Union[ec.Source, ec.Target], nrows: Optional[int] = None
):
    kwargs = {}
    if read_x:
        kwargs = read_x.model_dump(exclude_none=True)

    root_kwargs = (
        "nrows",
        "dtype",
        "sheet_name",
        "names",
        "encoding",
        "low_memory",
        "sep",
    )
    for k in root_kwargs:
        if hasattr(frame, k) and getattr(frame, k):
            kwargs[k] = getattr(frame, k)
    if nrows:
        kwargs["nrows"] = nrows

    return kwargs


def pull_frame(
    frame: Union[ec.Source, ec.Target],
    nrows: Optional[int] = None,
    # dtype=None,
) -> pd.DataFrame:
    if frame.type in ("mssql", "postgres", "duckdb"):
        df = pull_sql(frame)
    elif frame.type in (".csv", ".tsv"):
        if isinstance(frame, ec.Source):
            # kwargs = get_source_kwargs(frame.read_csv, nrows, dtype)
            kwargs = get_source_kwargs(frame.read_csv, frame, nrows)
            print(kwargs)
            if frame.type == ".tsv":
                kwargs["sep"] = "\t"
        else:
            kwargs = {}
        df = pull_csv(frame.file_path_dynamic, **kwargs)
    elif frame.type and frame.type in (".xlsx"):
        if isinstance(frame, ec.Source):
            # kwargs = get_source_kwargs(frame.read_excel, nrows, dtype)
            kwargs = get_source_kwargs(frame.read_excel, frame, nrows)
        else:
            kwargs = {}
        if frame.file_path in open_files:
            file = open_files[frame.file_path]
        else:
            file = frame.file_path_dynamic
        # kwargs["dtype"] = {1: "str"}
        df = pull_excel(file, **kwargs)
    elif frame.type in ("pandas"):
        if frame.table in staged_frames:
            df = staged_frames[frame.table]
        else:
            raise Exception("pandas frame not found")
    else:
        raise Exception("unable to build df")
    if isinstance(df.columns, pd.MultiIndex):
        if frame.stack:
            df = stack_columns(df, frame.stack)
        else:
            df = multiindex_to_singleindex(df)
    return pd.DataFrame(df)


def stack_columns(df, stack: ec.Stack):
    # Define the primary column headers based on the first four columns
    primary_headers = list(df.columns[: stack.fixed_columns])

    # Extract the top-level column names from the primary headers
    top_level_headers, _ = zip(*primary_headers)

    # Set the DataFrame's index to the primary headers
    df = df.set_index(primary_headers)

    # Get the names of the newly set indices
    current_index_names = list(df.index.names[: stack.fixed_columns])

    # Create a dictionary to map the current index names to the top-level headers
    index_name_mapping = dict(zip(current_index_names, top_level_headers))

    # Rename the indices using the created mapping
    df.index.rename(index_name_mapping, inplace=True)

    # Stack the DataFrame based on the top-level columns
    df = df.stack(level=stack.stack_header)

    # Rename the new index created by the stacking operation
    df.index.rename({None: stack.stack_name}, inplace=True)

    # Reset the index for the resulting DataFrame
    df.reset_index(inplace=True)

    return df


def multiindex_to_singleindex(df, separator="_"):
    df.columns = [separator.join(map(str, col)).strip() for col in df.columns.values]
    return df


def get_configs(config):
    target = config.target
    source = config.source
    add_cols = config.add_cols.model_dump()

    return target, source, add_cols


def add_columns(df: pd.DataFrame, add_cols: dict) -> pd.DataFrame:
    if add_cols:
        for k, v in add_cols.items():
            if (
                k != "additionalProperties"
                and v != ec.DynamicColumnValue.ROW_INDEX.value
            ):
                df[k] = v
    return df


def ingest(config: ec.Config) -> bool:
    target, source, add_cols = get_configs(config)
    consistent = frames_consistent(config)
    if (
        not target
        or not target.table
        or consistent
        or target.consistency == ec.TargetConsistencyValue.IGNORE.value
    ):
        # print(config.dtype)
        # source_df = get_df(source, config.nrows, config.dtype)
        source_df = pull_frame(source, config.nrows)
        source_df = add_columns(source_df, add_cols)
        return push_frame(source_df, target, add_cols)
    else:
        logging.error(target.table + ": Inconsistent, not saved.")
        return False


def build(config: ec.Config) -> bool:
    target, source, add_cols = get_configs(config)
    if (
        target
        and target.type not in ("pandas")
        and target.preparation_action != "no_action"
    ):
        action = target.preparation_action
        if action == "create_replace":
            df = pull_frame(source, 100)
            df = add_columns(df, add_cols)
            # res = build_sql_table(df, target, add_cols)
            res = build_target(df, target, add_cols)
        elif action == "truncate":
            res = truncate_target(target)
        elif action == "fail":
            logging.error("Table Exists, failing")
            res = False
        else:
            res = True
    else:
        res = True
    return res


def detect(config: ec.Config) -> bool:
    _, source, _ = get_configs(config)
    source = source.model_copy()
    source.nrows = 100

    df = pull_frame(source)
    print(df.dtypes.to_dict())
    return True


# def write_config(config: ec.Config) -> bool:
#     target, source, _ = get_configs(config)
