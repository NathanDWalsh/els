import yaml
from openpyxl import load_workbook
import logging
from typing import Optional, Union

from eel.path import ContentAwarePath as CAPath
import eel.config as ec


CONFIG_FILE_EXT = ".eel.yml"
FOLDER_CONFIG_FILE_STEM = "_"
ROOT_CONFIG_FILE_STEM = "__"
CONFIG_PREVIEW_FILE_NAME = "_preview.eel.yml"


def grow_branches(
    path: CAPath = CAPath(), parent: CAPath = None
) -> Union[CAPath, None]:
    if path.is_dir() and path.get_total_files() > 0:
        folder = CAPath(path.str, parent=parent)
        folder._config = get_paired_config(folder)
        ignore_configs = [
            path / get_folder_config_name(),
            path / get_root_config_name(),
            CAPath() / CONFIG_PREVIEW_FILE_NAME,
        ]
        for path_item in folder.iterdir():
            if not path_item.str.endswith(CONFIG_FILE_EXT):
                ignore_configs.append(CAPath(path_item.str + CONFIG_FILE_EXT))
                grow_branches(path_item, parent=folder)
            elif path_item not in ignore_configs:
                # TODO
                logging.error("ERROR: sole ymls not covered: " + path_item.str)
        return folder
    elif path.is_file() and not path.is_hidden():
        file = CAPath(path.str, parent=parent)
        file._config = get_paired_config(file)
        grow_leaves(file)
        return file
    return None


def grow_leaves(path: CAPath) -> None:
    if path.suffix == ".xlsx":
        sheet_names = get_sheet_names(path.str)
        for ws_name in sheet_names:
            content = CAPath(path.str, parent=path)
            content._config = get_paired_config(content, ws_name)
    else:
        content = CAPath((path / path.stem).str, parent=path)
        content._config = {}


def get_folder_config_name():
    return FOLDER_CONFIG_FILE_STEM + CONFIG_FILE_EXT


def get_root_config_name():
    return ROOT_CONFIG_FILE_STEM + CONFIG_FILE_EXT


def get_paired_config(item_path: CAPath, sub_path: str = ".") -> dict:
    config_path = get_paired_config_path(item_path)
    if config_path:
        ymls = get_yml_docs(config_path)
        configs = get_configs(ymls)
        for config, yml in zip(configs, ymls):
            if sub_path == config.sub_path:
                return yml
        return {}


def get_paired_config_path(path: CAPath) -> Optional[CAPath]:
    if path.is_root:
        config_path = path / get_root_config_name()
    elif path.is_dir():
        config_path = path / get_folder_config_name()
    elif path.is_file():
        if path.str.endswith(CONFIG_FILE_EXT):
            return path
        else:
            config_path = CAPath(path.str + CONFIG_FILE_EXT)
    if config_path.exists():
        return config_path
    return None


def merge_dicts_by_top_level_keys(*dicts: list[dict]) -> dict:
    merged_dict = {}
    for dict_ in dicts:
        for key, value in dict_.items():
            if (
                key in merged_dict
                and isinstance(value, dict)
                and (not merged_dict[key] is None)
            ):
                merged_dict[key].update(value)
            elif value is not None:
                # Add a new key-value pair to the merged dictionary
                merged_dict[key] = value
    return merged_dict


def get_sheet_names(file_path, sheet_states: list = ["visible"]):
    workbook = load_workbook(
        filename=file_path, read_only=True, data_only=True, keep_links=False
    )

    if sheet_states is None:
        worksheet_names = workbook.sheetnames
    else:
        worksheet_names = [
            sheet.title
            for sheet in workbook.worksheets
            if sheet.sheet_state in sheet_states
        ]

    workbook.close()
    return worksheet_names


def get_yml_docs(path: CAPath) -> list[dict]:
    with path.open() as file:
        yaml_text = file.read()
        documents = list(yaml.safe_load_all(yaml_text))
    return documents


def get_configs(ymls: list[dict]) -> list[ec.Config]:
    configs = []
    for yml in ymls:
        config = ec.Config(**yml)
        configs.append(config)
    return configs
